import cv2
import numpy as np
import openpyxl

from .transform_img import transform_img

def valid_data_loader(datadir, annotiondir):
    labeldir = annotiondir

    def reader(batch_size=50):
        images = []
        labels = []
        workbook = openpyxl.load_workbook(labeldir, data_only=True)
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            image = cv2.imread(datadir + '/' + row[1].value)
            image = transform_img(image)
            images.append(image)
            label = float(row[2].value)
            labels.append(label)
            if len(images) == batch_size:
                images_array = np.array(images).astype('float32')
                labels_array = np.array(labels).astype('float32').reshape(-1, 1)
                yield images_array, labels_array
                images = []
                labels = []
        if len(images) > 0:
            images_array = np.array(images).astype('float32')
            labels_array = np.array(labels).astype('float32').reshape(-1, 1)
            yield images_array, labels_array

    return reader
