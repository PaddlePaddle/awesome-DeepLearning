#!/usr/bin/env python
# coding: utf-8

# In[1]:


# 查看当前挂载的数据集目录, 该目录下的变更重启环境后会自动还原
# View dataset directory. 
# This directory will be recovered automatically after resetting environment. 
get_ipython().system('ls /home/aistudio/data')


# In[2]:


# 查看工作区文件, 该目录下的变更将会持久保存. 请及时清理不必要的文件, 避免加载过慢.
# View personal work directory. 
# All changes under this directory will be kept even after reset. 
# Please clean unnecessary files in time to speed up environment loading. 
# !ls /home/aistudio/work


# In[3]:


# 如果需要进行持久化安装, 需要使用持久化路径, 如下方代码示例:
# If a persistence installation is required, 
# you need to use the persistence path as the following: 
# !mkdir /home/aistudio/external-libraries
# !pip install beautifulsoup4 -t /home/aistudio/external-libraries


# In[4]:


# 同时添加如下代码, 这样每次环境(kernel)启动的时候只要运行下方代码即可: 
# Also add the following code, 
# so that every time the environment (kernel) starts, 
# just run the following code: 
# import sys 
# sys.path.append('/home/aistudio/external-libraries')


# In[ ]:





# 请点击[此处](https://ai.baidu.com/docs#/AIStudio_Project_Notebook/a38e5576)查看本环境基本用法.  <br>
# Please click [here ](https://ai.baidu.com/docs#/AIStudio_Project_Notebook/a38e5576) for more detailed instructions. 

# # **眼疾识别案例的ResNet实现**
# 本文参考paddle课程文档，通过paddle框架实现了ResNet模型基础B版本和变体D版本，通过模型API：version进行切换版本。训练部分和验证部分都采用分批训练的方式，用来解决模型迭代过程中显存不足的问题。 
# 
# **参考文献：**
# 
# > https://blog.csdn.net/sinat_17456165/article/details/106045728
# > 
# > https://aistudio.baidu.com/aistudio/education/preview/1533758
# > 
# > https://zhuanlan.zhihu.com/p/31852747/

# ### **数据集介绍**
# 
# 如今近视已经成为困扰人们健康的一项全球性负担，在近视人群中，有超过35%的人患有重度近视。近视会拉长眼睛的光轴，也可能引起视网膜或者络网膜的病变。随着近视度数的不断加深，高度近视有可能引发病理性病变，这将会导致以下几种症状：视网膜或者络网膜发生退化、视盘区域萎缩、漆裂样纹损害、Fuchs斑等。因此，及早发现近视患者眼睛的病变并采取治疗，显得非常重要。
# 
# `iChallenge-PM`是百度大脑和中山大学中山眼科中心联合举办的`iChallenge`比赛中，提供的关于病理性近视（Pathologic Myopia，PM）的医疗类数据集，包含1200个受试者的眼底视网膜图片，训练、验证和测试数据集各400张，其中训练集名称第一个字符表示类别，验证集的类别信息储存在PALM-Validation-GT的PM_Label_and_Fovea_Location.xlsx文件中。（以下图片仅供参考）
# 
# <center class="half">
#     <img src="https://raw.githubusercontent.com/buriedms/Eye-disease-recognition-ResNet-vd/main/eye_disease_recognition/images/data1.png" width="315"/>
#     <img src="https://github.com/buriedms/Eye-disease-recognition-ResNet-vd/blob/main/eye_disease_recognition/images/data2.png?raw=true" width="315"/>
#     <img src="https://github.com/buriedms/Eye-disease-recognition-ResNet-vd/blob/main/eye_disease_recognition/images/data3.png?raw=true" width="315"/>
# </center>
# 
# 图1  数据集的的大致情况

# ### **数据集的导入和预处理**

# ##### 导入相关库
# 导入相关的库，以备后续使用。

# In[5]:


import paddle
import paddle.nn as nn
import os
import cv2
import numpy as np
import openpyxl


# 数据集在`aistudio`平台上可直接载入数据集，并且通过以下代码指令，我们进行解压到指定位置。
# 
# 数据集data19469存放在data文件夹中

# In[6]:


if not os.path.isdir("train_data"):
    os.mkdir("train_data")
else:
    print('Train_data exist')
if not os.path.isdir('PALM-Training400'):
    get_ipython().system('unzip -oq /home/aistudio/data/data19469/training.zip')
    get_ipython().system('unzip -oq /home/aistudio/data/data19469/validation.zip')
    get_ipython().system('unzip -oq /home/aistudio/data/data19469/valid_gt.zip')
    get_ipython().system('unzip -oq /home/aistudio/PALM-Training400/PALM-Training400.zip -d /home/aistudio/train_data/')
else:
    print('The data has been decompressed')


# 基于cv2，对数据集进行导入,由于导入图像数据的格式为[H,W,C]，因此需要基于cv2.resize对结构进行重组为[C,H,W]。并且对数据进行相应的标准化操作。
# 
# **标准化作用 ：**
# 
# 1. 统一数据量纲
# 2. 平衡各特征的贡献
# 3. 加快了梯度下降求最优解的速度

# In[7]:


def transform_img(img):
    # 将图片尺寸缩放到 224x224
    img=cv2.resize(img,(224,224))
    # 读入的图像数据格式是[H,W,C]
    # 使用转置操作将其变成[C,H,W]
    img=np.transpose(img,(2,0,1))
    img.astype('float32')
    img=img/255.0
    img=img*2.0-1.0
    return img


# 通过自定义的`data_loader`和`valid_loader`导入训练集和验证集，并且在`data_loader`中打乱训练集。两者都预设`batch_size`选项设定每一个预设batch的大小。验证准确率和损失值由所有的batch的平均所得到。
# 
# 参数解释：
# 
# `datadir`：图片数据存在的文件夹路径
# 
# `annotiondir`：验证集标签文件路径
# 
# `batch_size`：每个批次的图片数据的数量
# 
# output：
# 
# 每个batch的图片数据，数据类型：`float32`，`numpy`保存，维度：[N,C,H,W]
# 
# **注意：其中训练集再导入时每个epoch都会进行随机打乱，而验证集不会**

# In[8]:


def data_loader(datadir,batch_size=10,mode='train'):
    filenames=os.listdir(datadir)
    def reader():
        if mode =='train':
            np.random.shuffle(filenames)
        batch_imgs=[]
        batch_labels=[]
        for name in filenames:
            filepath=os.path.join(datadir,name)
            img=cv2.imread(filepath)
            img=transform_img(img)
            if name[0]=='H' or name[0]=='N':
                label=0
            elif name[0]=='P':
                label=1
            elif name[0]=='V':
                continue
            else:
                raise('Not excepted file name')
            batch_imgs.append(img)
            batch_labels.append(label)
            if len(batch_imgs)==batch_size:
                imgs_array=np.array(batch_imgs).astype('float32')
                labels_array=np.array(batch_labels).astype('float32').reshape(-1,1)
                yield imgs_array,labels_array
                batch_imgs=[]
                batch_labels=[]
        if len(batch_imgs)>0:
            imgs_array=np.array(batch_imgs).astype('float32')
            labels_array=np.array(batch_labels).astype('float32').reshape(-1,1)
            yield imgs_array,labels_array
    return reader
    
def valid_data_loader(datadir,annotiondir):
    labeldir=annotiondir
    def reader(batch_size=50):
        images=[]
        labels=[]
        workbook=openpyxl.load_workbook(labeldir,data_only=True)
        worksheet=workbook.active
        for row in worksheet.iter_rows(min_row=2,max_row=worksheet.max_row):
            image=cv2.imread(datadir+'/'+row[1].value)
            image=transform_img(image)
            images.append(image)
            label=float(row[2].value)
            labels.append(label)
            if len(images)==batch_size:
                images_array=np.array(images).astype('float32')
                labels_array=np.array(labels).astype('float32').reshape(-1,1)
                yield images_array,labels_array
                images=[]
                labels=[]
        if len(images)>0:
            images_array=np.array(images).astype('float32')
            labels_array=np.array(labels).astype('float32').reshape(-1,1)
            yield images_array,labels_array
    return reader


# ### **`ResNet`模型及其变体D的理论解析**
# 
# ##### **`ResNet`模型背景**
# 
# 2015 年，`ResNet` 横空出世，一举斩获 CVPR 2016 最佳论文奖，而且在 `Imagenet` 比赛的三个任务以及 COCO 比赛的检测和分割任务上都获得了第一名。
# 
# 从经验来说，网络深度增加后，网络可以进行更加复杂的特征提取，因此可以取得更好的结果。但事实上并非如此，人们实验发现随着网络深度的增加，模型精度并不总是提升，并且这个问题显然不是由过拟合（overfitting）造成的，因为网络加深后不仅测试误差变高了，它的训练误差竟然也变高了。作者[何凯明](http://kaiminghe.com/)提出，这可能是因为更深的网络会伴随梯度消失/爆炸问题，从而阻碍网络的收敛。作者将这种加深网络深度但网络性能却下降的现象称为**退化问题**（degradation problem）。
# 
# ##### **`ResNet`的网络模型**
# 
# `ResNet`网络是参考了VGG19网络，在其基础上进行了修改，并通过短路机制加入了残差单元。
# 
# 变化主要体现在`ResNet`直接使用stride=2的卷积做下采样，并且用global average pool层替换了全连接层。`ResNet`的一个重要设计原则是：当feature map大小降低一半时，feature map的数量增加一倍，这保持了网络层的复杂度。以下为几种基本层数的`ResNet`版本的网络模型架构。
# 
# ![图2  `ResNet`的模型结构](https://github.com/buriedms/Eye-disease-recognition-ResNet-vd/blob/main/eye_disease_recognition/images/model_structure.png?raw=true)
# 
# 图2  `ResNet`的模型结构
# 
# 整个`ResNet`不使用`dropout`，全部使用BN。此外，回到最初的这张细节图，我们不难发现一些规律和特点：
# 
# - 全图大致分为5个模块，其中2-5模块是残差单元构成的模块
# - 受VGG的启发，卷积层主要是3×3卷积；
# 
# - **同一模块内图片的尺寸大小不变，不同模块之间相差大小减半，深度变为4倍**
# 
# - **第2个模块网络输出和输出图像尺寸相同，因此不需要下采样**
# - **第3-5模块的下采样仅操作一次，因此仅需要在每个模块的第一个`block`进行`stride=2`的下采样**
# 
# - 网络以平均池化层和`softmax`的全连接层结束，实际上工程上一般用自适应全局平均池化 (Adaptive Global Average Pooling)；

# ### **重点知识解读**
# 
# ##### **`Bottleneck`结构和1*1卷积**
# 
# **1x1卷积作用：**
# 
# - 对通道数进行升维和降维（跨通道信息整合），实现了多个特征图的线性组合，同时保持了原有的特征图大小；
# 
# - 相比于其他尺寸的卷积核，可以极大地降低运算复杂度；
# 
# - 如果使用两个3x3卷积堆叠，只有一个`relu`，但使用1x1卷积就会有两个`relu`，引入了更多的非线性映射；
# 
# ![图3  残差单元结构和1x1卷积](https://github.com/buriedms/Eye-disease-recognition-ResNet-vd/blob/main/eye_disease_recognition/images/bottleneck.png?raw=true)
# 
# 图3  残差单元结构和1x1卷积
# 
# （以上左图为`Basicblock`结构，右图为`Bottleneck`结构)
# 我们来计算一下1*1卷积的计算量优势：首先看上图右边的`bottleneck`结构，对于256维的输入特征，参数数目：
# $$
# 1*1*56*64+3*3*64*64+1*1*64*256=69632
# $$
# 如果同样的输入输出维度但不使用1x1卷积，而使用两个3x3卷积的话，参数数目:
# $$
# (3*3*256*256)*2=1179648
# $$
# 简单计算可知，使用了1x1卷积的`bottleneck`将计算量简化为原有的5.9%。
# 
# ##### **`ResNet-vb`及`ResNet-vd`模型设计**
# 
# 基于以上的规律和特点，我们做出如下设计：
# 
# 1. 为保证每个模块内部卷积前后的图像尺寸不变 ，将卷积**BN块的`padding`设计为`(kernel_size-1)//2`**，这就保证了`stride=1`图像尺寸不变，`stride=2`图像尺寸减半。
# 
# 2. 在2-5模块的残差单元block卷积采用如下类似的结构，**注意stride的设置**。
# 
# ![图4  `ResNet`的残差单元结构](https://github.com/buriedms/Eye-disease-recognition-ResNet-vd/blob/main/eye_disease_recognition/images/residual_unit.png?raw=true)
# 
# 图4  `ResNet`的残差单元结构
# 
# （以上左图为`ResNet-vb`，右图为`ResNet-vd`）
# 
# **注意：第2模块的stride=1，第3-5模块的stride=2实现下采样**

# ### **`ResNet`模型及变体的paddle实现**
# 参考paddle官方课程的ResNet网络模型教程搭建网络，并且修改为变体vd版本。
# 
# 详情参考[paddle课程文档](https://aistudio.baidu.com/aistudio/education/preview/1533758)。
# 

# In[9]:


# ResNet中使用BatchNorm层，在卷积层的后面加上BatchNorm以提升数值稳定性
# 定义卷积BN块
class ConvBNLayer(nn.Layer):
    def __init__(self,
                 num_channels,
                 num_filters,
                 filter_size,
                 stride=1,
                 groups=1,
                 act='relu'):
        """
        num_channels,卷积层的输入通道数
        num_filters,卷积层的输出通道数
        stride,卷积层的步幅
        groups,分组卷积的组数，默认groups=1不使用分组卷积
        """
        super(ConvBNLayer,self).__init__()
        self._conv=nn.Conv2D(
            in_channels=num_channels,
            out_channels=num_filters,
            kernel_size=filter_size,
            stride=stride,
            padding=(filter_size-1)//2,
            groups=groups,
            bias_attr=False,
        )
        self._batch_norm=nn.BatchNorm2D(num_filters)
        self.act=act
    def forward(self,inputs):
        x=self._conv(inputs)
        x=self._batch_norm(x)
        if self.act=='leaky':
            x=nn.functional.leaky_relu(x=x,negative_slope=0.1)
        elif self.act=='relu':
            x=nn.functional.relu(x=x)
        return x

# 定义残差块
# 每个残差块会对输入图片做三次卷积，然后跟输入图片进行短接
# 如果残差块中第三次卷积输出特征图的形状和输入不一致，则对输入图片做1x1卷积，将其输出形状调整为一致
class BottleneckBlock(nn.Layer):
    def __init__(self,
                 num_channels,
                 num_filters,
                 stride=1,
                 shortcut=True,
                 version='B'
                 ):
        super(BottleneckBlock,self).__init__()
        # 创建第一个1x1卷积层
        self.conv1=ConvBNLayer(
            num_channels=num_channels,
            num_filters=num_filters,
            filter_size=1,
            act='relu',
        )
        # 创建第二个3x3卷积层
        self.conv2=ConvBNLayer(
            num_channels=num_filters,
            num_filters=num_filters,
            filter_size=3,
            stride=stride,
            act='relu'
        )
        # 创建第三个1x1层，但是输出通道数乘4
        self.conv3=ConvBNLayer(
            num_channels=num_filters,
            num_filters=num_filters*4,
            filter_size=1,
            act='relu'
        )
        # 如果conv3的输出跟此残差块的输入数据形状一致，则shortcut=True
        # 否则shortcut=False，添加1个1x1的卷积作用在输入数据上，使其形状变为和conv3一样 
        if not shortcut:
            if version=='B':
                self.short=ConvBNLayer(
                    num_channels=num_channels,
                    num_filters=num_filters*4,
                    filter_size=1,
                    stride=stride
                )
            elif version=='D':
                self.short=nn.Sequential(
                    nn.AvgPool2D(kernel_size=stride,stride=stride),
                    ConvBNLayer(num_channels=num_channels,
                                num_filters=num_filters*4,
                                filter_size=1))      
            else:
                raise(f'bottleneck block version:{version} error, you can choice B or D')          
        self.shortcut=shortcut
        self._num_channels_out=num_filters*4
    def forward(self,inputs):
        x=self.conv1(inputs)
        x=self.conv2(x)
        x=self.conv3(x)
        # 如果shortcut=Ture，直接将inputs跟conv2的输出相加
        # 否则需要对inputs进行一次 卷积，将形状调整成跟conv2输出一致
        if self.shortcut:
            short=inputs
        else:
            short=self.short(inputs)
        y=paddle.add(x=short,y=x)
        return y

# 定义ResNet模型
class ResNet(nn.Layer):
    def __init__(self,layers=50,class_dim=10,version='B'):
        """
        layers,网络层数，可以可选项：50,101,152
        class_dim,分类标签的类别数
        """
        super(ResNet,self).__init__()
        self.version=version
        self.layers=layers
        supported_layers=[50,101,152]
        assert layers in supported_layers,        "supported layers are {} but input layer is {}".format(supported_layers,layers)
        # ResNet50包含的第2-5模块分别包括3,4,6,3个残差块
        if layers==50:
            depth=[3,4,6,3]
        # ResNet101包含的第2-5模块分别包括3,4,23,3个残差块
        if layers==101:
            depth=[3,4,23,3]
        # ResNet152包含的第2-5模块分别包括3,8,36,3个残差块
        if layers==152:
            depth=[3,8,36,3]
        # 第2-5模块所使用残差块的输出通道数
        num_filters=[64,128,256,512]

        # 第1个模块,64个7x7的卷积加上一个3x3最大化池化层，步长均为2
        self.conv=ConvBNLayer(
            num_channels=3,
            num_filters=64,
            filter_size=7,
            stride=2,
            act='relu')
        self.pool2d_max=nn.MaxPool2D(
            kernel_size=3,
            stride=2,
            padding=1)
        # 第2-5模块，使用各个残差块进行卷积操作
        self.bottleneck_block_list=[]
        num_channels=64
        for block in range(len(depth)):
            shortcut=False
            for i in range(depth[block]):
                bottleneck_block=self.add_sublayer(
                    'bb_%d_%d'%(block,i),
                    BottleneckBlock(
                        num_channels=num_channels,
                        num_filters=num_filters[block],
                        stride=2 if i==0 and block!=0 else 1,
                        shortcut=shortcut,
                        version=version))
                num_channels=bottleneck_block._num_channels_out
                self.bottleneck_block_list.append(bottleneck_block)
                shortcut=True

        # 在c5的输出特征图上使用全局池化
        self.pool2d_avg=nn.AdaptiveAvgPool2D(output_size=1)
        
        # stdv用来作为全连接层随机初始化参数的方差
        import math
        stdv=1.0/math.sqrt(2048*1.0)
        # 创建全连接层，输出大小为类别数目，经过残差网络的卷积核全局池化后，
        # 卷积特征的维度是[B,2048,1,1]，故最后一层全连接层的输入维度是2048
        self.out=nn.Linear(in_features=2048,out_features=class_dim,
        weight_attr=paddle.ParamAttr(
            initializer=paddle.nn.initializer.Uniform(-stdv,stdv)))
    
    def forward(self,inputs):
        x=self.conv(inputs)
        x=self.pool2d_max(x)
        for bottleneck_block in self.bottleneck_block_list:
            x=bottleneck_block(x)
        x=self.pool2d_avg(x)
        x=paddle.reshape(x,[x.shape[0],-1])
        x=self.out(x)
        return x


# ### 训练函数  
# 
# input:    
# `model`:待训练的模型  
# `datadir`:存放文件的主路径  
# `annotiondir`:存放标签数据的xlsx文件的路径    
# `optimizer`:优化模型参数所使用的优化器  
# `batch_size`：每个批次选取图片数量大小  
# `EPOCH_NUM`：训练的代数  
# `use_gpu`：是否使用GPU进行训练  
# `save`：模型保存的策略  
# 
# 相关代码参考如下。

# In[10]:


# 构建模型保存策略函数，返回为模型保存函数
def model_save(model_version):
    def save(save_model,model):
        if save_model:
            print('model save success !')
            if model==None:
                return 
            paddle.save(model.state_dict(),f'./model/resnet50_v{model_version}_PALM.pdparams')
            
    return save
save=model_save('C')
save(1>0,None)


# In[11]:


def train_pm(model,
             datadir,
             annotiondir,
             optimizer,
             batch_size=10,
             EPOCH_NUM=20,
             use_gpu=False,
             save=None):
    # 使用0号GPU训练
    paddle.set_device('gpu:0') if use_gpu else paddle.set_device('cpu')
    max_accuracy=0.

    print('********start training********')
    model.train()
    # 定义数据读取器
    train_loader=data_loader(datadir=datadir+'/train_data/PALM-Training400',batch_size=batch_size,mode='train')
    valid_loader=valid_data_loader(datadir+'/PALM-Validation400',annotiondir)
    for epoch in range(EPOCH_NUM):
        for batch_id,data in enumerate(train_loader()):
            x_data,y_data=data
            img=paddle.to_tensor(x_data)
            label=paddle.to_tensor(y_data).astype('int64')
            # 使用模型进行前向计算，得到预测值
            out=model(img)
            loss=nn.functional.cross_entropy(out,label,reduction='none')
            avg_loss=paddle.mean(loss)
            if batch_id%10==0:
                print("epoch:{}===batch_id:{}===loss:{:.4f}".format(
                    epoch,batch_id,float(avg_loss.numpy())))
            # 反向传播，更新权重，消除梯度
            avg_loss.backward()
            optimizer.step()
            optimizer.clear_grad()
        valid_accuracy,valid_loss=valid_pm(model,valid_loader,batch_size=50)
        print('[validation]:======accuracy:{:.5f}/loss:{:.5f}'.format(valid_accuracy,valid_loss))
        if save!=None and valid_accuracy>max_accuracy:
            save(valid_accuracy>max_accuracy,model)
            max_accuracy=valid_accuracy
            print('max accuracy :',max_accuracy)
        print()

        


# ###  验证函数
# 
# input:  
# 
# `model`:待训练的模型  
# `valid_loader`:验证数据的迭代生成器  
# `batch_size`:每一个批次验证数据的大小  
# 
# 
# 相关代码参考如下。
# 
# **注意：为保证避免显存问题，采用分批次验证，求平均值**

# In[12]:


def valid_pm(model,valid_loader,batch_size=100):
    model.eval()
    print("*****valid data import success*****")
    batch_accuracy=[]
    batch_loss=[]
    for batch_id,data in enumerate(valid_loader(batch_size=batch_size)):
        x_data,y_data=data
        img=paddle.to_tensor(x_data)
        label=paddle.to_tensor(y_data).astype('int64')
        out=model(img)
        predict=paddle.argmax(out,1)
        loss=nn.functional.cross_entropy(out,label,reduction='none')
        avg_loss=paddle.mean(loss)
        accuracy=sum(predict.numpy().reshape(-1,1)==label.numpy())/float(label.shape[0])
        batch_loss.append(float(avg_loss.numpy()))
        batch_accuracy.append(accuracy)
        # print('batch_id:{}===accuracy:{}/loss:{}'.format(batch_id,accuracy,avg_loss.numpy()))
        # if batch_id==1:
        #     print('predict:{}'.format(predict.numpy()))
        #     print('label  :{}'.format(label.numpy()[:,0]))
    avg_loss=np.mean(batch_loss)
    avg_accuracy=np.mean(batch_accuracy)
    return avg_accuracy,avg_loss
        


# ### 超参数及训练部分
# 超参数含义：
# 
# `model_version`：选择使用的ResNet版本，可选B或D，默认B；  
# `use_gpu`：是否使用gpu进行训练，默认True；  
# `lr`：学习率；  
# `momentum`：动量系数  
# `load_model`：是否载入预训练模型，默认True   
# `save_model`：是否保存训练模型，默认False
# 
# **注释：因为在2.1.2的更新中线下了paddle.save对paddle.nn.Layer的支持，所以更换模型保存方式**
# 
# **后期优化想法：根据验证集中正确率是否上升来确定本次的模型参数是否保存，达到选取最优参数的目的。**(over)

# In[13]:


filedir=os.getcwd()
model_version='D'
use_gpu=True
lr=0.001
momentum=0.3
load_model=True
save_model=False

model=ResNet(layers=50,version=model_version)
if os.path.exists(f'./model/resnet50_v{model_version}_PALM.pdmodel') and load_model:
    model_params=paddle.load(f'./model/resnet50_v{model_version}_PALM.pdparams')
    model.set_state_dict(model_params)
annotion_path=filedir+'/PALM-Validation-GT/PM_Label_and_Fovea_Location.xlsx'
optimizer=paddle.optimizer.Momentum(learning_rate=lr,momentum=momentum,parameters=model.parameters())

print('文件主路径：',filedir)
print('训练模型版本：',model_version)
print('是否采用预训练模型：',load_model)
print('是否采用GPU：',use_gpu)

if save_model:
    save=model_save(model_version)
else:
    save=None
train_pm(model,filedir,annotion_path,optimizer,use_gpu=use_gpu,save=save)

# paddle.save(model,f'./model/resnet50_v{model_version}_PALM.pdmodel')


# ### **ResNet-vb导入和验证**

# In[14]:


annotion_path='./PALM-Validation-GT/PM_Label_and_Fovea_Location.xlsx'
valid_loader=valid_data_loader('./PALM-Validation400',annotion_path)
model_version='B'
model=ResNet(layers=50,version=model_version)

model_params=paddle.load(f'./model/resnet50_v{model_version}_PALM.pdparams')
model.set_state_dict(model_params)
valid_accuracy,valid_loss=valid_pm(model,valid_loader,batch_size=50)
print('[validation]:======accuracy:{:.5f}/loss:{:.5f}'.format(valid_accuracy,valid_loss))


# ### **ResNet-vd的导入和验证**

# In[15]:


annotion_path='./PALM-Validation-GT/PM_Label_and_Fovea_Location.xlsx'
valid_loader=valid_data_loader('./PALM-Validation400',annotion_path)
model_version='D'
model=ResNet(layers=50,version=model_version)

model_params=paddle.load(f'./model/resnet50_v{model_version}_PALM.pdparams')
model.set_state_dict(model_params)
valid_accuracy,valid_loss=valid_pm(model,valid_loader,batch_size=50)
print('[validation]:======accuracy:{:.5f}/loss:{:.5f}'.format(valid_accuracy,valid_loss))


# ### **总结**
# 
# 本次实验通过眼疾识别案例，来进一步的理解在aistudio平台下导入训练集和进行训练集的预处理操作，这其中包含了很多图像预处理方面的基础知识和基本操作，更有体现了平台所特有的在线操作文件特性，让代码编辑工作不再局限于个人电脑的配置，有利于广大代码爱好者的学习和进一步提升。后续使用paddle框架进一步全面阐述`ResNet`网络的基本原理和及其变体D版本的特性，在并且在基础上基于框架将其实现，这其中包含大量的网络模型图的阅读和理解、1x1卷积的巧妙作用和模型的巧妙设计等等知识，值得深层次的体会和学习。
# 
# 在此次实验中，生成器的使用和分批次的训练和验证这是一个非常必要且有效的节省显存的办法，这也体现了在运行大型项目工程的实际当中，代码的参数量、运行效率和占用空间是必须要考虑的因素。
