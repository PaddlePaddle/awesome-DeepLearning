#!/usr/bin/env python
# coding: utf-8

# In[1]:


# 查看当前挂载的数据集目录, 该目录下的变更重启环境后会自动还原
# View dataset directory. 
# This directory will be recovered automatically after resetting environment. 
get_ipython().system('ls /home/aistudio/data')


# In[2]:


# 查看工作区文件, 该目录下的变更将会持久保存. 请及时清理不必要的文件, 避免加载过慢.
# View personal work directory. 
# All changes under this directory will be kept even after reset. 
# Please clean unnecessary files in time to speed up environment loading. 
# !ls /home/aistudio/work


# In[3]:


# 如果需要进行持久化安装, 需要使用持久化路径, 如下方代码示例:
# If a persistence installation is required, 
# you need to use the persistence path as the following: 
# !mkdir /home/aistudio/external-libraries
# !pip install beautifulsoup4 -t /home/aistudio/external-libraries


# In[4]:


# 同时添加如下代码, 这样每次环境(kernel)启动的时候只要运行下方代码即可: 
# Also add the following code, 
# so that every time the environment (kernel) starts, 
# just run the following code: 
# import sys 
# sys.path.append('/home/aistudio/external-libraries')


# 请点击[此处](https://ai.baidu.com/docs#/AIStudio_Project_Notebook/a38e5576)查看本环境基本用法.  <br>
# Please click [here ](https://ai.baidu.com/docs#/AIStudio_Project_Notebook/a38e5576) for more detailed instructions. 

# # 眼疾识别案例的ResNet实现
# 本文通过paddle框架实现了ResNet模型基础B版本和变体D版本，通过模型API：version进行切换版本。训练部分和验证部分都采用分批训练的方式，用来解决模型迭代过程中显存不足的问题。   

# In[2]:


import paddle
import paddle.nn as nn
import os
import cv2
import numpy as np
import openpyxl


# ### 数据的解压缩

# In[5]:


if not os.path.isdir("train_data"):
    os.mkdir("train_data")
else:
    print('Train_data exist')
if not os.path.isdir('PALM-Training400'):
    get_ipython().system('unzip -oq /home/aistudio/data/data19469/training.zip')
    get_ipython().system('unzip -oq /home/aistudio/data/data19469/validation.zip')
    get_ipython().system('unzip -oq /home/aistudio/data/data19469/valid_gt.zip')
    get_ipython().system('unzip -oq /home/aistudio/PALM-Training400/PALM-Training400.zip -d /home/aistudio/train_data/')
else:
    print('The data has been decompressed')


# ### 图像的预处理

# In[ ]:


def transform_img(img):
    # 将图片尺寸缩放到 224x224
    img=cv2.resize(img,(224,224))
    # 读入的图像数据格式是[H,W,C]
    # 使用转置操作将其变成[C,H,W]
    img=np.transpose(img,(2,0,1))
    img.astype('float32')
    img=img/255.0
    img=img*2.0-1.0
    return img


# ### 训练数据和验证数据的读入

# In[7]:


def data_loader(datadir,batch_size=10,mode='train'):
    filenames=os.listdir(datadir)
    def reader():
        if mode =='train':
            np.random.shuffle(filenames)
        batch_imgs=[]
        batch_labels=[]
        for name in filenames:
            filepath=os.path.join(datadir,name)
            img=cv2.imread(filepath)
            img=transform_img(img)
            if name[0]=='H' or name[0]=='N':
                label=0
            elif name[0]=='P':
                label=1
            elif name[0]=='V':
                continue
            else:
                raise('Not excepted file name')
            batch_imgs.append(img)
            batch_labels.append(label)
            if len(batch_imgs)==batch_size:
                imgs_array=np.array(batch_imgs).astype('float32')
                labels_array=np.array(batch_labels).astype('float32').reshape(-1,1)
                yield imgs_array,labels_array
                batch_imgs=[]
                batch_labels=[]
        if len(batch_imgs)>0:
            imgs_array=np.array(batch_imgs).astype('float32')
            labels_array=np.array(batch_labels).astype('float32').reshape(-1,1)
            yield imgs_array,labels_array
    return reader
def valid_data_loader(datadir,annotiondir):
    labeldir=annotiondir
    def reader(batch_size=50):
        images=[]
        labels=[]
        workbook=openpyxl.load_workbook(labeldir,data_only=True)
        worksheet=workbook.active
        for row in worksheet.iter_rows(min_row=2,max_row=worksheet.max_row):
            image=cv2.imread(datadir+'/'+row[1].value)
            image=transform_img(image)
            images.append(image)
            label=float(row[2].value)
            labels.append(label)
            if len(images)==batch_size:
                images_array=np.array(images).astype('float32')
                labels_array=np.array(labels).astype('float32').reshape(-1,1)
                yield images_array,labels_array
                images=[]
                labels=[]
        if len(images)>0:
            images_array=np.array(images).astype('float32')
            labels_array=np.array(labels).astype('float32').reshape(-1,1)
            yield images_array,labels_array
    return reader


# ### ResNet网络模型

# In[8]:


# ResNet中使用BatchNorm层，在卷积层的后面加上BatchNorm以提升数值稳定性
# 定义卷积BN块
class ConvBNLayer(nn.Layer):
    def __init__(self,
                 num_channels,
                 num_filters,
                 filter_size,
                 stride=1,
                 groups=1,
                 act='relu'):
        """
        num_channels,卷积层的输入通道数
        num_filters,卷积层的输出通道数
        stride,卷积层的步幅
        groups,分组卷积的组数，默认groups=1不使用分组卷积
        """
        super(ConvBNLayer,self).__init__()
        self._conv=nn.Conv2D(
            in_channels=num_channels,
            out_channels=num_filters,
            kernel_size=filter_size,
            stride=stride,
            padding=(filter_size-1)//2,
            groups=groups,
            bias_attr=False,
        )
        self._batch_norm=nn.BatchNorm2D(num_filters)
        self.act=act
    def forward(self,inputs):
        x=self._conv(inputs)
        x=self._batch_norm(x)
        if self.act=='leaky':
            x=nn.functional.leaky_relu(x=x,negative_slope=0.1)
        elif self.act=='relu':
            x=nn.functional.relu(x=x)
        return x

# 定义残差块
# 每个残差块会对输入图片做三次卷积，然后跟输入图片进行短接
# 如果残差块中第三次卷积输出特征图的形状和输入不一致，则对输入图片做1x1卷积，将其输出形状调整为一致
class BottleneckBlock(nn.Layer):
    def __init__(self,
                 num_channels,
                 num_filters,
                 stride=1,
                 shortcut=True,
                 version='B'
                 ):
        super(BottleneckBlock,self).__init__()
        # 创建第一个1x1卷积层
        self.conv1=ConvBNLayer(
            num_channels=num_channels,
            num_filters=num_filters,
            filter_size=1,
            act='relu',
        )
        # 创建第二个3x3卷积层
        self.conv2=ConvBNLayer(
            num_channels=num_filters,
            num_filters=num_filters,
            filter_size=3,
            stride=stride,
            act='relu'
        )
        # 创建第三个1x1层，但是输出通道数乘4
        self.conv3=ConvBNLayer(
            num_channels=num_filters,
            num_filters=num_filters*4,
            filter_size=1,
            act='relu'
        )
        # 如果conv3的输出跟此残差块的输入数据形状一致，则shortcut=True
        # 否则shortcut=False，添加1个1x1的卷积作用在输入数据上，使其形状变为和conv3一样 
        if not shortcut:
            if version=='B':
                self.short=ConvBNLayer(
                    num_channels=num_channels,
                    num_filters=num_filters*4,
                    filter_size=1,
                    stride=stride
                )
            elif version=='D':
                self.short=nn.Sequential(
                    nn.AvgPool2D(kernel_size=stride,stride=stride),
                    ConvBNLayer(num_channels=num_channels,
                                num_filters=num_filters*4,
                                filter_size=1))      
            else:
                raise(f'bottleneck block version:{version} error, you can choice B or D')          
        self.shortcut=shortcut
        self._num_channels_out=num_filters*4
    def forward(self,inputs):
        x=self.conv1(inputs)
        x=self.conv2(x)
        x=self.conv3(x)
        # 如果shortcut=Ture，直接将inputs跟conv2的输出相加
        # 否则需要对inputs进行一次 卷积，将形状调整成跟conv2输出一致
        if self.shortcut:
            short=inputs
        else:
            short=self.short(inputs)
        y=paddle.add(x=short,y=x)
        return y

# 定义ResNet模型
class ResNet(nn.Layer):
    def __init__(self,layers=50,class_dim=10,version='B'):
        """
        layers,网络层数，可以可选项：50,101,152
        class_dim,分类标签的类别数
        """
        super(ResNet,self).__init__()
        self.version=version
        self.layers=layers
        supported_layers=[50,101,152]
        assert layers in supported_layers,        "supported layers are {} but input layer is {}".format(supported_layers,layers)
        # ResNet50包含的第2-5模块分别包括3,4,6,3个残差块
        if layers==50:
            depth=[3,4,6,3]
        # ResNet101包含的第2-5模块分别包括3,4,23,3个残差块
        if layers==101:
            depth=[3,4,23,3]
        # ResNet152包含的第2-5模块分别包括3,8,36,3个残差块
        if layers==152:
            depth=[3,8,36,3]
        # 第2-5模块所使用残差块的输出通道数
        num_filters=[64,128,256,512]

        # 第1个模块,64个7x7的卷积加上一个3x3最大化池化层，步长均为2
        self.conv=ConvBNLayer(
            num_channels=3,
            num_filters=64,
            filter_size=7,
            stride=2,
            act='relu')
        self.pool2d_max=nn.MaxPool2D(
            kernel_size=3,
            stride=2,
            padding=1)
        # 第2-5模块，使用各个残差块进行卷积操作
        self.bottleneck_block_list=[]
        num_channels=64
        for block in range(len(depth)):
            shortcut=False
            for i in range(depth[block]):
                bottleneck_block=self.add_sublayer(
                    'bb_%d_%d'%(block,i),
                    BottleneckBlock(
                        num_channels=num_channels,
                        num_filters=num_filters[block],
                        stride=2 if i==0 and block!=0 else 1,
                        shortcut=shortcut,
                        version=version))
                num_channels=bottleneck_block._num_channels_out
                self.bottleneck_block_list.append(bottleneck_block)
                shortcut=True

        # 在c5的输出特征图上使用全局池化
        self.pool2d_avg=nn.AdaptiveAvgPool2D(output_size=1)
        
        # stdv用来作为全连接层随机初始化参数的方差
        import math
        stdv=1.0/math.sqrt(2048*1.0)
        # 创建全连接层，输出大小为类别数目，经过残差网络的卷积核全局池化后，
        # 卷积特征的维度是[B,2048,1,1]，故最后一层全连接层的输入维度是2048
        self.out=nn.Linear(in_features=2048,out_features=class_dim,
        weight_attr=paddle.ParamAttr(
            initializer=paddle.nn.initializer.Uniform(-stdv,stdv)))
    
    def forward(self,inputs):
        x=self.conv(inputs)
        x=self.pool2d_max(x)
        for bottleneck_block in self.bottleneck_block_list:
            x=bottleneck_block(x)
        x=self.pool2d_avg(x)
        x=paddle.reshape(x,[x.shape[0],-1])
        x=self.out(x)
        return x


# ### ResNet的变体 ResNet-vd

# In[9]:


# 定义残差块BottleneckBlock_vd模块
# 每个残差块会对输入图片做三次卷积，然后跟输入图片进行短接
# 如果残差块中第三次卷积输出特征图的形状和输入不一致，则对输入图片做1x1卷积，将其输出形状调整为一致
class BottleneckBlock_vd(nn.Layer):
    def __init__(self,
                 num_channels,
                 num_filters,
                 stride=1,
                 shortcut=True,
                 ):
        super(BottleneckBlock_vd,self).__init__()
        # 创建第一个1x1卷积层
        self.conv1=ConvBNLayer(
            num_channels=num_channels,
            num_filters=num_filters,
            filter_size=1,
            act='relu',
        )
        # 创建第二个3x3卷积层
        self.conv2=ConvBNLayer(
            num_channels=num_filters,
            num_filters=num_filters,
            filter_size=3,
            stride=stride,
            act='relu'
        )
        # 创建第三个1x1层，但是输出通道数乘4
        self.conv3=ConvBNLayer(
            num_channels=num_filters,
            num_filters=num_filters*4,
            filter_size=1,
            act='relu'
        )
        # 如果conv3的输出跟此残差块的输入数据形状一致，则shortcut=True
        # 否则shortcut=False，添加1个1x1的卷积作用在输入数据上，使其形状变为和conv3一样 
        if not shortcut:
            self.short=nn.Sequential(
                nn.AvgPool2D(kernel_size=stride,stride=stride),
                ConvBNLayer(num_channels=num_channels,
                            num_filters=num_filters*4,
                            filter_size=1))
        self.shortcut=shortcut
        self._num_channels_out=num_filters*4
    def forward(self,inputs):
        x=self.conv1(inputs)
        x=self.conv2(x)
        x=self.conv3(x)
        # 如果shortcut=Ture，直接将inputs跟conv2的输出相加
        # 否则需要对inputs进行一次卷积，将形状调整成跟conv2输出一致
        if self.shortcut:
            short=inputs
        else:
            # print('inputs:',inputs.shape)
            short=self.short(inputs)
            # print('output:',short.shape)
        y=paddle.add(x=short,y=x)
        return y

# 定义ResNet_vd模型
class ResNet_vd(nn.Layer):
    def __init__(self,layers=50,class_dim=10):
        """
        layers,网络层数，可以可选项：50,101,152
        class_dim,分类标签的类别数
        """
        super(ResNet_vd,self).__init__()
        self.layers=layers
        supported_layers=[50,101,152]
        assert layers in supported_layers,        "supported layers are {} but input layer is {}".format(supported_layers,layers)
        # ResNet50包含的第2-5模块分别包括3,4,6,3个残差块
        if layers==50:
            depth=[3,4,6,3]
        # ResNet101包含的第2-5模块分别包括3,4,23,3个残差块
        if layers==101:
            depth=[3,4,23,3]
        # ResNet152包含的第2-5模块分别包括3,8,36,3个残差块
        if layers==152:
            depth=[3,8,36,3]
        # 第2-5模块所使用残差块的输出通道数
        num_filters=[64,128,256,512]

        # 第1个模块,64个7x7的卷积加上一个3x3最大化池化层，步长均为2
        self.conv=ConvBNLayer(
            num_channels=3,
            num_filters=64,
            filter_size=7,
            stride=2,
            act='relu')
        self.pool2d_max=nn.MaxPool2D(
            kernel_size=3,
            stride=2,
            padding=1)
        # 第2-5模块，使用各个残差块进行卷积操作
        self.bottleneck_block_vd_list=[]
        num_channels=64
        for block in range(len(depth)):
            shortcut=False
            for i in range(depth[block]):
                bottleneck_block_vd=self.add_sublayer(
                    'bb_%d_%d'%(block,i),
                    BottleneckBlock_vd(
                        num_channels=num_channels,
                        num_filters=num_filters[block],
                        stride=2 if i==0 and block!=0 else 1,
                        shortcut=shortcut))
                num_channels=bottleneck_block_vd._num_channels_out
                self.bottleneck_block_vd_list.append(bottleneck_block_vd)
                shortcut=True

        # 在c5的输出特征图上使用全局池化
        self.pool2d_avg=nn.AdaptiveAvgPool2D(output_size=1)
        
        # stdv用来作为全连接层随机初始化参数的方差
        import math
        stdv=1.0/math.sqrt(2048*1.0)
        # 创建全连接层，输出大小为类别数目，经过残差网络的卷积核全局池化后，
        # 卷积特征的维度是[B,2048,1,1]，故最后一层全连接层的输入维度是2048
        self.out=nn.Linear(in_features=2048,out_features=class_dim,
        weight_attr=paddle.ParamAttr(
            initializer=paddle.nn.initializer.Uniform(-stdv,stdv)))
    
    def forward(self,inputs):
        # print('0:',inputs.shape)
        x=self.conv(inputs)
        # print('1:',x.shape)
        x=self.pool2d_max(x)
        # print('2:',x.shape)
        for bottleneck_block in self.bottleneck_block_vd_list:
            x=bottleneck_block(x)
            # print(3+self.bottleneck_block_vd_list.index(bottleneck_block),':',x.shape)
        x=self.pool2d_avg(x)
        x=paddle.reshape(x,[x.shape[0],-1])
        x=self.out(x)
        return x


# ###  验证函数
# 分批次验证，求平均值

# In[10]:


def valid_pm(model,valid_loader,batch_size=100):
    model.eval()
    print("*****valid data import success*****")
    batch_accuracy=[]
    batch_loss=[]
    for batch_id,data in enumerate(valid_loader(batch_size=batch_size)):
        x_data,y_data=data
        img=paddle.to_tensor(x_data)
        label=paddle.to_tensor(y_data).astype('int64')
        out=model(img)
        predict=paddle.argmax(out,1)
        loss=nn.functional.cross_entropy(out,label,reduction='none')
        avg_loss=paddle.mean(loss)
        accuracy=sum(predict.numpy().reshape(-1,1)==label.numpy())/float(label.shape[0])
        batch_loss.append(float(avg_loss.numpy()))
        batch_accuracy.append(accuracy)
        # print('batch_id:{}===accuracy:{}/loss:{}'.format(batch_id,accuracy,avg_loss.numpy()))
        # if batch_id==1:
        #     print('predict:{}'.format(predict.numpy()))
        #     print('label  :{}'.format(label.numpy()[:,0]))
    avg_loss=np.mean(batch_loss)
    avg_accuracy=np.mean(batch_accuracy)
    return avg_accuracy,avg_loss
        


# ### 训练函数
# 采用gpu进行训练

# In[11]:


def train_pm(model,
             datadir,
             annotiondir,
             optimizer,
             batch_size=10,
             EPOCH_NUM=20,
             use_gpu=False):
    # 使用0号GPU训练
    paddle.set_device('gpu:0') if use_gpu else paddle.set_device('cpu')

    print('********start training********')
    model.train()
    # 定义数据读取器
    train_loader=data_loader(datadir=datadir+'/train_data/PALM-Training400',batch_size=batch_size,mode='train')
    valid_loader=valid_data_loader(datadir+'/PALM-Validation400',annotiondir)
    for epoch in range(EPOCH_NUM):
        for batch_id,data in enumerate(train_loader()):
            x_data,y_data=data
            img=paddle.to_tensor(x_data)
            label=paddle.to_tensor(y_data).astype('int64')
            # 使用模型进行前向计算，得到预测值
            out=model(img)
            loss=nn.functional.cross_entropy(out,label,reduction='none')
            avg_loss=paddle.mean(loss)
            if batch_id%10==0:
                print("epoch:{}===batch_id:{}===loss:{:.4f}".format(
                    epoch,batch_id,float(avg_loss.numpy())))
            # 反向传播，更新权重，消除梯度
            avg_loss.backward()
            optimizer.step()
            optimizer.clear_grad()
        valid_accuracy,valid_loss=valid_pm(model,valid_loader,batch_size=50)
        print('[validation]:======accuracy:{:.5f}/loss:{:.5f}'.format(valid_accuracy,valid_loss))


# ### 训练的超参数及训练部分

# In[12]:


filedir=os.getcwd()
print('文件主路径：',filedir)
model_version='D'
lr=0.001
load_model=False
if os.path.exists(f'./model/resnet50_v{model_version}_PALM.pdmodel') and load_model:
    model=paddle.load(f'./model/resnet50_v{model_version}_PALM.pdmodel')
else:
    model=ResNet(layers=50,version=model_version)
annotion_path=filedir+'/PALM-Validation-GT/PM_Label_and_Fovea_Location.xlsx'
optimizer=paddle.optimizer.Momentum(learning_rate=lr,momentum=0.9,parameters=model.parameters())
use_gpu=True
train_pm(model,filedir,annotion_path,optimizer,use_gpu=use_gpu)
paddle.save(model,f'./model/resnet50_v{model_version}_PALM.pdmodel_1')


# ### ResNet导入和验证

# In[13]:


annotion_path='./PALM-Validation-GT/PM_Label_and_Fovea_Location.xlsx'
valid_loader=valid_data_loader('./PALM-Validation400',annotion_path)
model_version='B'
model=paddle.load(f'./model/resnet50_v{model_version}_PALM.pdmodel')
valid_accuracy,valid_loss=valid_pm(model,valid_loader,batch_size=50)
print('[validation]:======accuracy:{:.5f}/loss:{:.5f}'.format(valid_accuracy,valid_loss))


# ### ResNet-vd的导入和验证

# In[14]:


annotion_path='./PALM-Validation-GT/PM_Label_and_Fovea_Location.xlsx'
valid_loader=valid_data_loader('./PALM-Validation400',annotion_path)
model_version='D'
model=paddle.load(f'./model/resnet50_v{model_version}_PALM.pdmodel')
valid_accuracy,valid_loss=valid_pm(model,valid_loader,batch_size=50)
print('[validation]:======accuracy:{:.5f}/loss:{:.5f}'.format(valid_accuracy,valid_loss))

